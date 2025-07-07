__author__ = "Edgar Nandayapa"
__version__ = "v0.0.1 2023"

from glob import glob
import pandas as pd
import seaborn as sns
import operator
import os
import re
import numpy as np
#import matplotlib
#import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
# from openpyxl.styles import Font
import warnings
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px

#matplotlib.use('module://ipympl.backend_nbagg')

warnings.filterwarnings("ignore", message=".*cannot be placed.*", category=UserWarning)
def is_running_in_jupyter():
    """Check if code is running in Jupyter notebook"""
    try:
        from IPython import get_ipython
        return get_ipython() is not None
    except ImportError:
        return False

import plotly.io as pio
pio.renderers.default = 'notebook' if is_running_in_jupyter() else 'browser'


def find_and_list_files(folder_path):
    file_patterns = ["JV_*.txt", "**/*JV_*.csv", "**/*JV_*.txt"]
    file_list = []
    for pattern in file_patterns:
        file_list.extend(glob(os.path.join(folder_path, pattern), recursive=True))
    file_list = list(set(file_list))

    return file_list


def load_files(file_list):
    # Consolidate file patterns for searching
    file_list.sort(key=natural_keys)
    # file_list = sorted(file_list, key=lambda x: int(x.split('\\')[-1].replace('JV_', '').replace('.txt', '')))

    # Initialize empty DataFrames for merged results
    jv_chars_merged = pd.DataFrame()
    curves_merged = pd.DataFrame()

    # Process each file
    for file_path in file_list:
        try:
            # Extract JV Characteristics and JV curve from the file
            jv_chars, jv_curve = process_file(file_path)
            # Merge data into cumulative DataFrames
            jv_chars_merged = pd.concat([jv_chars_merged, jv_chars],
                                        ignore_index=True) if not jv_chars.empty else jv_chars_merged
            curves_merged = pd.concat([curves_merged, jv_curve]) if not jv_curve.empty else curves_merged

        except Exception as e:  # Catch all exceptions to avoid stopping the loop
            print(f"Error processing {file_path}: {e}")
    curves_merged = curves_merged.reset_index()
    # Check if data was successfully loaded
    if jv_chars_merged.empty and curves_merged.empty:
        print("One of the files has an issue.")

    return jv_chars_merged, curves_merged


def replace_current_density_unit(idx):
    # This regular expression matches (mA/cm²) or (mA/cm^2) and captures the "mA/cm" part before the ² or ^2
    pattern = r'\(mA/cm(?:²|\^2)\)'
    replacement = '(mA/cm2)'
    return re.sub(pattern, replacement, idx)


def process_file(file_path):
    # Determines delimiter based on file extension
    linepos = find_separators_in_file(file_path)
    delimiter = '\t' if file_path.endswith('.txt') else ','

    try:
        # Initial attempt to read JV Characteristics
        jv_chars = pd.read_csv(file_path, skiprows=linepos[0], header=0, index_col=0, nrows=9,
                               delimiter=delimiter).transpose()
        # Attempt to read JV Curve - adjust parameters as per your file structure
        jv_curve = pd.read_csv(file_path, skiprows=linepos[1], header=0, index_col=None,
                               delimiter=delimiter).transpose()

        # Replace problematic character
        jv_chars.columns = [col.replace('²', '2') for col in jv_chars.columns]
        jv_curve.index = [replace_current_density_unit(idx) for idx in jv_curve.index]

        if not jv_chars.empty:
            jv_chars = add_extra_info(jv_chars, file_path, data_type='chars')

        if not jv_curve.empty:
            jv_curve = add_extra_info(jv_curve, file_path, data_type='curve')

    except pd.errors.EmptyDataError:
        jv_chars = pd.DataFrame()
        jv_curve = pd.DataFrame()

    return jv_chars, jv_curve


def add_extra_info(df, file_path, data_type):
    """
    Adds extra information to the DataFrame based on file path and data type.

    Parameters:
    - df: DataFrame to augment.
    - file_path: Path of the file being processed.
    - data_type: Type of data ('chars' for JV characteristics, 'curve' for JV curve).

    Returns:
    - DataFrame with added information.
    """
    norm_path = os.path.normpath(file_path)
    df['sample'] = file_path.split("JV_")[-1].rsplit(".", 1)[0]
    df['batch'] = norm_path.split(os.sep)[-2]
    df['condition'] = pd.NA

    split_index = df.index.to_series().str.split('_', expand=True)
    if data_type == "chars":
        df[['cell', 'direction', 'ilum']] = split_index

    if data_type == 'curve':
        # Assign the split results to new columns in df
        df[['variable', 'cell', 'direction', 'ilum']] = split_index

    return df


def find_separators_in_file(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    positions = []
    for index, line in enumerate(lines):
        if line.strip() == "--":
            positions.append(index + 1)
            # print(f"'--' found at line {index + 1}")

    return positions


def atoi(text):
    return int(text) if text.isdigit() else text


def natural_keys(text):
    return [atoi(c) for c in re.split(r'(\d+)', text)]


def name_by_condition(data, key_list, value_list):
    condition_dict = dict(zip(key_list, value_list))

    data["condition"] = data["sample"].map(condition_dict)

    return data


def data_filter_setup(df, filter_list):
    # Filter conditions
    # par = ["PCE(%)", "FF(%)", "FF(%)", "Voc(V)", "Jsc(mA/cm2)", "ilum"]
    # ope = ["<", "<", ">", "<", ">", "=="]
    # val = [40, 89, 24, 2, -30, "Light"]
    if not filter_list:
        filter_list = [("PCE(%)", "<", "40"), ("FF(%)", "<", "89"), ("FF(%)", ">", "24"), ("Voc(V)", "<", "2"),
                       ("Jsc(mA/cm2)", ">", "-30")]

    # List of operators
    operat = {"<": operator.lt, ">": operator.gt, "==": operator.eq,
              "<=": operator.le, ">=": operator.ge, "!=": operator.ne}

    data = df.copy()

    # Initialize the filter_reason column with empty strings
    data['filter_reason'] = ''
    filtering_options = []
    # for col, op, va in zip(par, ope, val):
    for col, op, va in filter_list:
        # Update the filter_reason for rows that do not meet the condition
        mask = operat[op](data[col], float(va))
        data.loc[~mask, 'filter_reason'] += f'{col} {op} {va}, '
        filtering_options.append(f'{col} {op} {va}')

    # Filter out rows that have any filter_reason
    trash = data[data['filter_reason'] != ''].copy()
    # Remove rows from data that were moved to trash
    data = data[data['filter_reason'] == '']
    # Clean up the filter_reason string by removing the trailing comma and space
    trash['filter_reason'] = trash['filter_reason'].str.rstrip(', ')

    print(f"\n {trash.shape[0]} of {df.shape[0]} samples were removed based on the specified filters: "
          f"{',  '.join(filtering_options)}.\n")
    print(trash[['sample', 'cell', 'filter_reason']].to_string(index=False))

    return data, trash, filtering_options


def jv_plot_curve_best(path, jvc, cur):
    """Plot the JV curve of the best device using Plotly"""
    # Find best device
    index_num = jvc["PCE(%)"].idxmax()
    sample = jvc.loc[index_num]["sample"]
    cell = jvc.loc[index_num]["cell"]

    # Filter data to focus on best device
    focus = cur.loc[(cur["sample"] == sample) & (cur["cell"] == cell)]

    if len(focus) == 0:
        sample = jvc.loc[index_num]["identifier"]
        focus = cur.loc[(cur["sample"] == sample) & (cur["cell"] == cell)]

    plotted = focus.copy().drop(
        columns=["index", "sample", "cell", "direction", "ilum", "batch", "condition"]).set_index(["variable"]).T
    dire = focus.loc[(focus["variable"] == "Voltage (V)")]["direction"].values
    ilum = focus.loc[(focus["variable"] == "Voltage (V)")]["ilum"].values

    # Create Plotly figure
    fig = go.Figure()

    # Add x and y axis lines at 0
    fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, line=dict(color="gray", width=2))
    fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, line=dict(color="gray", width=2))

    # Plot each direction
    for c, p in enumerate(dire):
        x = plotted["Voltage (V)"].iloc[:, c]
        y = plotted["Current Density(mA/cm2)"].iloc[:, c]

        marker_symbol = 'x' if dire[c] == "Reverse" else 'circle'
        
        fig.add_trace(go.Scatter(
            x=x, 
            y=y,
            mode='lines+markers',
            marker=dict(symbol=marker_symbol),
            name=f"{dire[c]} ({ilum[c]})",
            hovertemplate='Voltage: %{x:.3f} V<br>Current Density: %{y:.3f} mA/cm²<br>%{text}',
            text=[f"Sample: {sample}, Cell: {cell}" for _ in x]
        ))

    # Get JV characteristics values
    df_rev = jvc.loc[(jvc["sample"] == sample) & (jvc["cell"] == cell) & (jvc["direction"] == "Reverse")]
    df_for = jvc.loc[(jvc["sample"] == sample) & (jvc["cell"] == cell) & (jvc["direction"] == "Forward")]

    # Extract values
    char_vals = ['Voc(V)', 'Jsc(mA/cm2)', 'FF(%)', 'PCE(%)']
    char_rev = []
    char_for = []
    
    for cv in char_vals:
        char_rev.append(df_rev[cv].values[0])
        char_for.append(df_for[cv].values[0])

    # Add MPP points
    v_f = df_for['V_mpp(V)'].values[0]
    v_r = df_rev['V_mpp(V)'].values[0]
    j_f = df_for['J_mpp(mA/cm2)'].values[0]
    j_r = df_rev['J_mpp(mA/cm2)'].values[0]

    fig.add_trace(go.Scatter(
        x=[v_f], y=[j_f],
        mode='markers',
        marker=dict(color='red', size=10),
        name='Forward MPP',
        hoverinfo='text',
        hovertext=f'MPP Forward<br>V: {v_f:.3f} V<br>J: {j_f:.3f} mA/cm²'
    ))

    fig.add_trace(go.Scatter(
        x=[v_r], y=[j_r],
        mode='markers',
        marker=dict(color='red', size=10, symbol='x'),
        name='Reverse MPP',
        hoverinfo='text',
        hovertext=f'MPP Reverse<br>V: {v_r:.3f} V<br>J: {j_r:.3f} mA/cm²'
    ))

    # Add JV information as annotations
    text_rev = f"""        Rev:
<br>Voc: {char_rev[0]:>5.2f}
<br>Jsc:  {char_rev[1]:>5.1f}
<br>FF:   {char_rev[2]:>5.1f}
<br>PCE: {char_rev[3]:>5.1f}"""
    text_for = f"For:<br>{char_for[0]:.2f} V<br>{char_for[1]:.1f} mA/cm²<br>{char_for[2]:.1f}%<br>{char_for[3]:.1f}%"

    # Add annotations for values
    fig.add_annotation(
        x=0.24, y=-5,
        text=text_rev,
        showarrow=False,
        font=dict(size=12),
        align="left"
    )
    
    fig.add_annotation(
        x=0.55, y=-5,
        text=text_for,
        showarrow=False,
        font=dict(size=12),
        align="left"
    )

    # Add sample name annotation
    fig.add_annotation(
        x=0.20, y=1.5,
        #xref="paper", yref="paper",
        text=f"Sample: {sample} ({cell})",
        showarrow=False,
        font=dict(size=13),
        align="left",
    )

    # Update layout
    fig.update_layout(
        title="JV Curve - Best Device",
        xaxis_title='Voltage [V]',
        yaxis_title='Current Density [mA/cm²]',
        xaxis=dict(range=[-0.2, 1.35]),
        yaxis=dict(range=[-25, 3]),
        template="plotly_white",
        legend=dict(
            x=1.02,
            y=1,
            xanchor="left",
            yanchor="top"
        ),
        grid=dict(rows=1, columns=1)
    )

    # Add grid lines
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')

    # Save figure if not in Jupyter
    sample_name = "JV_best_device.html"
    if not is_running_in_jupyter():
        fig.write_html(path + sample_name)
        print(f"Saved JV curve of best device")

    return fig, sample_name



def jv_plot_by_cell_3x2(df, sample, path):
    """Plot JV curves for each cell in a 2x3 grid using Plotly"""
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    
    # Filter the DataFrame for the specified sample
    focus = df[df["sample"] == sample]

    # Group the DataFrame by 'cell'
    grouped = focus.groupby('cell')
    
    # Create 2x3 subplot grid
    fig = make_subplots(rows=2, cols=3, 
                        subplot_titles=[f"Cell {cell}" for cell in grouped.groups.keys()],
                        shared_xaxes=True, shared_yaxes=True)

    # Dynamically identify measurement columns
    # These are numeric columns that aren't metadata
    metadata_cols = ["index", "sample", "cell", "direction", "ilum", "batch", "condition", "variable"]
    all_cols = df.columns.tolist()
    measurement_cols = [col for col in all_cols if col not in metadata_cols and col != '']
    
    print(f"Debug: Available columns: {all_cols}")
    print(f"Debug: Metadata columns: {metadata_cols}")
    print(f"Debug: Identified measurement columns: {measurement_cols}")
    
    # Iterate through each cell
    for i, (cell, group) in enumerate(grouped):
        row = i // 3 + 1  # Calculate row (1-indexed)
        col = i % 3 + 1   # Calculate column (1-indexed)
        
        if i >= 6:  # Check to prevent index error if there are more than 6 cells
            break

        # Add x and y axis lines at 0
        fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, 
                    line=dict(color="gray", width=2), row=row, col=col)
        fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, 
                    line=dict(color="gray", width=2), row=row, col=col)

        # Process each direction and illumination combination
        for direction in group['direction'].unique():
            for ilum in group['ilum'].unique():
                # Get voltage data
                voltage_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Voltage (V)")
                ]
                
                # Get current density data
                current_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Current Density(mA/cm2)")
                ]
                
                if voltage_row.empty or current_row.empty:
                    print(f"Debug: No data for Cell {cell}, {direction}, {ilum}")
                    continue
                
                print(f"Debug: Processing Cell {cell}, {direction}, {ilum}")
                print(f"Debug: Voltage row shape: {voltage_row.shape}")
                print(f"Debug: Current row shape: {current_row.shape}")
                
                # Extract the measurement values using available columns
                try:
                    voltage_values = voltage_row[measurement_cols].iloc[0].values
                    current_values = current_row[measurement_cols].iloc[0].values
                    
                    print(f"Debug: Extracted {len(voltage_values)} voltage points")
                    print(f"Debug: Extracted {len(current_values)} current points")
                    
                    # Convert to mA/cm² if values are very small (likely in A/cm²)
                    if abs(current_values.max()) < 0.1:  # If max current is less than 0.1, likely in A
                        current_values_ma = current_values * 1000  # Convert A to mA
                        print(f"Debug: Converted to mA/cm² (factor 1000)")
                    else:
                        current_values_ma = current_values
                        print(f"Debug: Using values as-is (already in mA/cm²)")
                    
                    # Plot the curve
                    label = f"{direction} ({ilum})"
                    
                    fig.add_trace(
                        go.Scatter(
                            x=voltage_values, 
                            y=current_values_ma,
                            mode="lines+markers",
                            name=label,
                            legendgroup=label,
                            showlegend=(i == 0),  # Only show legend for the first cell
                            hovertemplate='Voltage: %{x:.3f} V<br>Current Density: %{y:.3f} mA/cm²<br>Cell: ' + str(cell) + '<br>' + label
                        ),
                        row=row, col=col
                    )
                    print(f"Debug: Successfully added trace for {label}")
                    
                except Exception as e:
                    print(f"Debug: Error processing Cell {cell}, {direction}, {ilum}: {e}")
                    print(f"Debug: Voltage row columns: {voltage_row.columns.tolist()}")
                    continue

    # Update layout
    fig.update_layout(
        title=f"Sample {sample}",
        height=700,
        width=1000,
        template="plotly_white",
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5
        )
    )

    # Update all x and y axes
    for i in range(1, 7):
        if i <= len(grouped):
            row = (i-1) // 3 + 1
            col = (i-1) % 3 + 1
            
            fig.update_xaxes(
                title_text="Voltage [V]" if row == 2 else None,
                range=[-0.2, 1.35],
                showgrid=True, 
                gridwidth=1, 
                gridcolor='lightgray',
                row=row, col=col
            )
            
            fig.update_yaxes(
                title_text="Current Density [mA/cm²]" if col == 1 else None,
                range=[-25, 3],
                showgrid=True, 
                gridwidth=1, 
                gridcolor='lightgray',
                row=row, col=col
            )

    # Save figure if not in Jupyter
    image_name = f"JV_cells_by_sample_{sample}.html"
    if not is_running_in_jupyter():
        fig.write_html(path + image_name)
        print(f"Saved JV_cells_by_sample_{sample}.html")

    return fig, image_name


def jv_plot_by_substrate(df, sample, path):
    """Plot JV curves for all cells on a substrate using Plotly"""
    import plotly.graph_objects as go
    
    focus = df[df["sample"] == sample]
    
    # Create Plotly figure
    fig = go.Figure()

    # Add x and y axis lines at 0
    fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, line=dict(color="gray", width=2))
    fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, line=dict(color="gray", width=2))

    grouped = focus.groupby('cell')
    
    # Create a color scale for different cells
    colors = [f"hsl({i*360/grouped.ngroups},100%,50%)" for i in range(grouped.ngroups)]
    
    # Dynamically identify measurement columns
    metadata_cols = ["index", "sample", "cell", "direction", "ilum", "batch", "condition", "variable"]
    all_cols = df.columns.tolist()
    measurement_cols = [col for col in all_cols if col not in metadata_cols and col != '']
    
    print(f"Debug: Available columns: {all_cols}")
    print(f"Debug: Identified measurement columns: {measurement_cols}")

    for i, ((cell, group), color) in enumerate(zip(grouped, colors)):
        # Iterate through directions and illumination conditions
        for direction in group['direction'].unique():
            for ilum in group['ilum'].unique():
                # Get voltage data
                voltage_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Voltage (V)")
                ]
                
                # Get current density data
                current_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Current Density(mA/cm2)")
                ]
                
                if voltage_row.empty or current_row.empty:
                    print(f"Debug: No data for Cell {cell}, {direction}, {ilum}")
                    continue
                
                print(f"Debug: Processing Cell {cell}, {direction}, {ilum}")
                
                # Extract the measurement values using available columns
                try:
                    voltage_values = voltage_row[measurement_cols].iloc[0].values
                    current_values = current_row[measurement_cols].iloc[0].values
                    
                    # Convert to mA/cm² if values are very small (likely in A/cm²)
                    if abs(current_values.max()) < 0.1:  # If max current is less than 0.1, likely in A
                        current_values_ma = current_values * 1000  # Convert A to mA
                    else:
                        current_values_ma = current_values

                    # Determine line style based on direction
                    line_dash = 'solid' if direction == 'Forward' else 'dash'
                    
                    # Plot the curve
                    label = f"Cell {cell}, {direction} ({ilum})"
                    
                    fig.add_trace(go.Scatter(
                        x=voltage_values, 
                        y=current_values_ma,
                        mode="lines",
                        name=label,
                        line=dict(color=color, dash=line_dash),
                        hovertemplate='Voltage: %{x:.3f} V<br>Current Density: %{y:.3f} mA/cm²<br>' + label
                    ))
                    print(f"Debug: Successfully added trace for {label}")
                    
                except Exception as e:
                    print(f"Debug: Error processing Cell {cell}, {direction}, {ilum}: {e}")
                    print(f"Debug: Available columns in voltage_row: {voltage_row.columns.tolist()}")
                    continue

    # Update layout
    fig.update_layout(
        title=f"Sample {sample}",
        xaxis_title='Voltage [V]',
        yaxis_title='Current Density [mA/cm²]',
        xaxis=dict(range=[-0.2, 1.35]),
        yaxis=dict(range=[-25, 3]),
        template="plotly_white",
        legend=dict(
            x=1.02,
            y=1,
            xanchor="left",
            yanchor="top"
        )
    )

    # Add grid lines
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')

    # Save figure if not in Jupyter
    image_name = f"JV_combined_sample_{sample}.html"
    if not is_running_in_jupyter():
        fig.write_html(path + image_name)
        print(f"Saved JV_combined_sample_{sample}.html")

    return fig, image_name


def jv_plot_together(df1, df2, path, namestring):
    """Plot all JV curves together using Plotly with full interactivity and individual legends"""
    # Prepare the data frame as before
    if namestring == "All":
        df2_copy = df2.copy()
    else:
        # For "Filtered" and "Omitted" modes, only use curves from samples/cells in df1
        # Get the unique sample/cell combinations from df1
        sample_cells = df1[['sample', 'cell']].drop_duplicates()
        
        # Filter df2 to only include these sample/cell combinations
        df2_copy = pd.merge(sample_cells, df2, on=['sample', 'cell'], how='inner')[df2.columns]

    df2_plot = drop_extra_cols_and_ready_to_plot(df2_copy)

    cols = []
    counters = {'Voltage (V)': 0, 'Current Density(mA/cm2)': 0}
    for col in df2_plot.columns:
        counters[col] += 1
        cols.append(f"{col} {counters[col]}")
    df2_plot.columns = cols

    # Create Plotly figure
    fig = go.Figure()

    # Add x and y axis lines at 0
    fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, line=dict(color="gray", width=2))
    fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, line=dict(color="gray", width=2))

    # Extract relevant information about the source data
    source_data = df2_copy[['sample', 'cell', 'direction', 'variable']].drop_duplicates()
    source_data = source_data[source_data['variable'] == 'Voltage (V)'] if 'variable' in source_data.columns else source_data
    
    # Get sample/cell combinations - these will be our unique traces
    unique_samples = source_data[['sample', 'cell']].drop_duplicates()
    
    # Create a color palette with enough colors for all curves
    import plotly.express as px
    import random
    import colorsys
    
    # Generate a large set of distinct colors
    def generate_distinct_colors(n):
        colors = []
        for i in range(n):
            # Use golden ratio to space hues evenly
            hue = (i * 0.618033988749895) % 1
            # Use middle-high saturation and value for visibility
            saturation = 0.7 + random.random() * 0.3
            value = 0.7 + random.random() * 0.3
            # Convert HSV to RGB
            r, g, b = colorsys.hsv_to_rgb(hue, saturation, value)
            # Convert to hex
            color = f"rgb({int(r*255)}, {int(g*255)}, {int(b*255)})"
            colors.append(color)
        return colors
    
    # Get more colors than we need just to be safe
    num_colors_needed = len(unique_samples) * 2  # For forward and reverse
    all_colors = generate_distinct_colors(num_colors_needed)
    random.shuffle(all_colors)  # Shuffle to avoid sequential colors being too similar
    
    # Get PCE values for all samples if available
    pce_values = {}
    if 'PCE(%)' in df1.columns:
        for _, row in df1.iterrows():
            key = (row['sample'], row['cell'], row['direction'])
            pce_values[key] = row['PCE(%)']
    
    # Populate traces for all samples with distinct colors
    added_traces = {}  # To keep track of what we've already added
    
    # Add each curve with its own color and legend entry
    num_pairs = len(df2_plot.columns) // 2
    for i in range(1, num_pairs + 1):
        if i <= len(source_data) and not df2_plot[f'Voltage (V) {i}'].empty and not df2_plot[f'Current Density(mA/cm2) {i}'].empty:
            # Get information about this curve
            if i <= len(source_data):
                sample_info = source_data.iloc[i-1]
                sample_name = sample_info['sample'] if 'sample' in sample_info else f"Sample {i}"
                cell_name = sample_info['cell'] if 'cell' in sample_info else f"Cell {i}"
                direction = sample_info['direction'] if 'direction' in sample_info else "Unknown"
            else:
                sample_name = f"Sample {i}"
                cell_name = f"Cell {i}"
                direction = "Unknown"
            
            # Get the PCE value if available
            pce_val = pce_values.get((sample_name, cell_name, direction), None)
            pce_str = f", PCE: {pce_val:.2f}%" if pce_val is not None else ""
            
            # Create a unique identifier for this sample/cell combination
            sample_cell_key = (sample_name, cell_name)
            
            # Assign a consistent color to this sample/cell combination
            if sample_cell_key not in added_traces:
                # First time seeing this sample/cell, assign a color
                color_idx = len(added_traces) % len(all_colors)
                added_traces[sample_cell_key] = {
                    'color': all_colors[color_idx], 
                    'forward_added': False,
                    'reverse_added': False
                }
            
            # Set the line style based on scan direction
            is_forward = direction == 'Forward' if isinstance(direction, str) else False
            line_dash = 'solid' if is_forward else 'dash'
            
            # Check if we've already added this direction
            if is_forward and added_traces[sample_cell_key]['forward_added']:
                continue
            if not is_forward and added_traces[sample_cell_key]['reverse_added']:
                continue
            
            # Update our tracking
            if is_forward:
                added_traces[sample_cell_key]['forward_added'] = True
            else:
                added_traces[sample_cell_key]['reverse_added'] = True
            
            # Create a descriptive name
            curve_name = f"Sample {sample_name} Cell {cell_name} ({direction}{pce_str})"
            
            # Add trace with hover info and legend entry
            fig.add_trace(go.Scatter(
                x=df2_plot[f'Voltage (V) {i}'],
                y=df2_plot[f'Current Density(mA/cm2) {i}'],
                mode='lines',
                line=dict(
                    color=added_traces[sample_cell_key]['color'], 
                    width=2, 
                    dash=line_dash
                ),
                name=curve_name,
                legendgroup=f"sample_{sample_name}_cell_{cell_name}",
                showlegend=True,  # Show in legend
                hovertemplate=(
                    "Voltage: %{x:.3f} V<br>" +
                    "Current Density: %{y:.3f} mA/cm²<br>" +
                    f"Sample: {sample_name}<br>" +
                    f"Cell: {cell_name}<br>" +
                    f"Direction: {direction}{pce_str}" +
                    "<extra></extra>"
                )
            ))
            
    # Update layout with improved organization for many legend entries
    fig.update_layout(
        title={
            'text': f'{namestring} J-V Curves<br><sup>Interactive plot with {len(added_traces)} sample-cell combinations</sup>',
            'y': 0.95,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top',
            'font': dict(size=18)
        },
        xaxis_title={
            'text': 'Voltage [V]',
            'font': dict(size=14)
        },
        yaxis_title={
            'text': 'Current Density [mA/cm²]',
            'font': dict(size=14)
        },
        xaxis=dict(range=[-0.2, 1.35]),
        yaxis=dict(range=[-25, 3]),
        template="plotly_white",
        # Setup legend to handle many entries
        legend={
            'x': 1.02,
            'y': 1,
            'xanchor': 'left',
            'yanchor': 'top',
            'title': {'text': 'Samples & Cells'},
            'bordercolor': 'rgba(0,0,0,0.2)',
            'borderwidth': 1,
            'itemsizing': 'constant',  # Make legend items the same size
            'font': {'size': 10},  # Smaller font for legend entries
            'tracegroupgap': 5  # Reduce gap between legend groups
        },
        plot_bgcolor='rgb(250,250,250)',
        margin=dict(l=80, r=200, t=100, b=80),  # Extended right margin for legend
        height=700,  # Taller plot to accommodate legend
    )

    # Add grid lines
    fig.update_xaxes(
        showgrid=True, 
        gridwidth=1, 
        gridcolor='lightgray',
        zeroline=True,
        zerolinecolor='rgba(0,0,0,0.3)',
        zerolinewidth=1
    )
    
    fig.update_yaxes(
        showgrid=True, 
        gridwidth=1, 
        gridcolor='lightgray',
        zeroline=True,
        zerolinecolor='rgba(0,0,0,0.3)',
        zerolinewidth=1
    )

    # Add annotation explaining how to use the legend
    fig.add_annotation(
        x=1.0, 
        y=-0.15,
        xref="paper", 
        yref="paper",
        text="Click on legend items to show/hide individual curves",
        showarrow=False,
        font=dict(size=12, color="gray"),
        align="right"
    )

    # Save figure if not in Jupyter
    image_name = f"JV_together_{namestring}.html"
    if not is_running_in_jupyter():
        fig.write_html(path + image_name)
        print(f"Saved JV_together_{namestring}.html")

    return fig, image_name


def drop_extra_cols_and_ready_to_plot(df):
    cols_to_remove = ["index", "sample", "cell", "direction", "ilum", "batch", "condition"]

    common_cols_to_remove = df.columns.intersection(cols_to_remove)
    df_clean = df.copy().drop(columns=common_cols_to_remove)

    df_clean = df_clean.set_index(["variable"]).T
    return df_clean


def boxplot_all_cells(path, wb, data, var_x, var_y, filtered_info, datatype):
    """Create a boxplot with all cells using Plotly with much wider boxes"""
    names_dict = {
        "voc": 'Voc(V)', "jsc": 'Jsc(mA/cm2)', "ff": 'FF(%)', "pce": 'PCE(%)',
        "vmpp": 'V_mpp(V)', "jmpp": 'J_mpp(mA/cm2)', "pmpp": 'P_mpp(mW/cm2)',
        "rser": 'R_series(Ohmcm2)', "rshu": 'R_shunt(Ohmcm2)'
    }
    var_name_y = names_dict[var_y]
    trash, filters = filtered_info

    try:
        data["sample"] = data["sample"].astype(int)
    except ValueError:
        pass

    data['Jsc(mA/cm2)'] = data['Jsc(mA/cm2)'].abs()

    # Calculate statistics 
    descriptor = data.groupby(var_x)[var_name_y].describe()

    # Ordering
    order_parameter = "alphabetic"
    if order_parameter != "alphabetic":
        orderc = descriptor.sort_values(by=[order_parameter])["count"].index
    else:
        orderc = descriptor.sort_index()["count"].index

    # Create dictionaries to map categories to their counts
    data_counts = data.groupby(var_x)[var_name_y].count().to_dict()
    trash_counts = trash.groupby(var_x)[var_name_y].count().to_dict() if not trash.empty else {}

    # Create figure
    fig = go.Figure()
    
    # Use a pleasing color palette
    colors = [
        'rgba(93, 164, 214, 0.7)', 'rgba(255, 144, 14, 0.7)', 
        'rgba(44, 160, 101, 0.7)', 'rgba(255, 65, 54, 0.7)', 
        'rgba(207, 114, 255, 0.7)', 'rgba(127, 96, 0, 0.7)',
        'rgba(255, 140, 184, 0.7)', 'rgba(79, 90, 117, 0.7)',
        'rgba(222, 158, 54, 0.7)', 'rgba(82, 182, 133, 0.7)',
        'rgba(148, 103, 189, 0.7)', 'rgba(23, 190, 207, 0.7)'
    ]
    
    # Add each category's boxplot
    for i, category in enumerate(orderc):
        # Get data for this category
        category_data = data[data[var_x] == category][var_name_y].dropna()
        if not category_data.empty:
            # Get counts and statistics
            data_count = data_counts.get(category, 0)
            trash_count = trash_counts.get(category, 0)
            median = category_data.median()
            mean = category_data.mean()
            
            # Format category name with count
            category_name = f"{category} (n={data_count})" if trash_count == 0 else f"{category} ({data_count}/{data_count + trash_count})"
            
            # Add boxplot with improved styling - much wider boxes
            fig.add_trace(go.Box(
                y=category_data,
                name=category_name,
                boxpoints='all',     # Show all points
                pointpos=0,          # Center points horizontally
                jitter=0.5,          # Add jitter to points
                whiskerwidth=0.4,    # Thicker whiskers
                marker=dict(
                    size=5,          # Slightly larger points
                    opacity=0.7,     # Semi-transparent
                    color='rgba(0,0,0,0.7)'  # Points in dark gray
                ),
                line=dict(width=1.5),  # Slightly thicker lines
                fillcolor=colors[i % len(colors)],
                boxmean=True,        # Show mean
                hoverinfo='all',
                hovertemplate=(
                    f"<b>{category}</b><br>" +
                    "Value: %{y:.3f}<br>" +
                    f"Median: {median:.3f}<br>" +
                    f"Mean: {mean:.3f}<br>" +
                    f"Count: {data_count}"
                ),
                # Add notches for visual appeal
                notched=False,
                notchwidth=0.5
            ))
    
    # Create a title with data information
    title_text = f"Boxplot of {var_y} by {var_x}" + (" (filtered out)" if datatype == "junk" else "")
    subtitle = f"Data from {len(data)} ({trash.shape[0]} removed) measurements across {data[var_x].nunique()} {var_x} categories"
    
    # Update layout with MUCH more compressed spacing
    fig.update_layout(
        title={
            'text': f"{title_text}<br><sup>{subtitle}</sup>",
            'y': 0.95,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top',
            'font': dict(size=18)
        },
        xaxis_title=var_x,
        yaxis_title=var_name_y,
        # These are the key settings to make boxes much wider
        boxmode='group',
        boxgap=0.05,         # Very minimal gap between boxes (5% of box width)
        boxgroupgap=0.1,     # Very minimal gap between groups (10% of box width)
        # Force the width of bars/boxes to be much larger
        bargap=0.05,         # Almost no gap between bars
        bargroupgap=0.1,     # Almost no gap between bar groups
        template="plotly_white",
        margin=dict(l=40, r=40, t=100, b=80),
        showlegend=False,
        plot_bgcolor='rgb(243, 243, 243)',  # Light gray background
        paper_bgcolor='rgb(243, 243, 243)',
    )
    
    # Force the box width using custom width ratio
    fig.update_traces(
        width=0.8,            # Make boxes very wide (80% of available space)
        quartilemethod="linear"  # Use linear method for quartiles
    )
    
    # Rotate x-axis labels if many categories
    if len(orderc) > 4:
        fig.update_layout(
            plot_bgcolor='white',    # Sets the plot area background to white
            paper_bgcolor='white',    # Sets the entire figure background to white
            xaxis=dict(
                tickangle=-10,
                tickfont=dict(size=10)
            )
        )
    
    # Add annotation about removed samples
    #if trash.shape[0] > 0:
    #    fig.add_annotation(
    #        x=1, y=1,
    #        xref="paper", yref="paper",
    #        text=f"Removed: {trash.shape[0]} samples",
    #        showarrow=False,
    #        font=dict(size=12, color='rgba(0,0,0,0.7)'),
    #        align="right",
    #        bgcolor="rgba(255,255,255,0.8)",
    #        bordercolor="rgba(0,0,0,0.2)",
    #        borderwidth=1,
    #        borderpad=4
    #    )
    
    # Save to Excel
    wb = save_combined_excel_data(path, wb, data, filtered_info, var_x, var_name_y, var_y, descriptor)

    if datatype == "junk":
        sample_name = f"boxplotj_{var_y}_by_{var_x}.html"
    else:
        sample_name = f"boxplot_{var_y}_by_{var_x}.html"

    if not is_running_in_jupyter():
        fig.write_html(f"{path}{sample_name}")
        print(f"Saved boxplot of {var_y} by {var_x}")

    return fig, sample_name, wb


def boxplot_paired_by_direction(path, wb, data, var_x, var_y, filtered_info, datatype):
    """Create a paired boxplot with forward/backward direction using different colors"""
    names_dict = {
        "voc": 'Voc(V)', "jsc": 'Jsc(mA/cm2)', "ff": 'FF(%)', "pce": 'PCE(%)',
        "vmpp": 'V_mpp(V)', "jmpp": 'J_mpp(mA/cm2)', "pmpp": 'P_mpp(mW/cm2)',
        "rser": 'R_series(Ohmcm2)', "rshu": 'R_shunt(Ohmcm2)'
    }
    var_name_y = names_dict[var_y]
    trash, filters = filtered_info

    try:
        data["sample"] = data["sample"].astype(int)
    except ValueError:
        pass

    data['Jsc(mA/cm2)'] = data['Jsc(mA/cm2)'].abs()

    # Check if 'direction' column exists
    if 'direction' not in data.columns:
        print("Warning: 'direction' column not found in data. Creating dummy direction column.")
        data['direction'] = 'forward'  # Default fallback

    # Calculate statistics for each direction
    descriptor = data.groupby([var_x, 'direction'])[var_name_y].describe()

    # Ordering based on var_x categories
    order_parameter = "alphabetic"
    if order_parameter != "alphabetic":
        orderc = data.groupby(var_x)[var_name_y].describe().sort_values(by=[order_parameter])["count"].index
    else:
        orderc = data.groupby(var_x)[var_name_y].describe().sort_index()["count"].index

    # Create dictionaries to map categories and directions to their counts
    data_counts = data.groupby([var_x, 'direction'])[var_name_y].count().to_dict()
    trash_counts = trash.groupby([var_x, 'direction'])[var_name_y].count().to_dict() if not trash.empty and 'direction' in trash.columns else {}

    # Create figure
    fig = go.Figure()
    
    # Define colors for forward and backward directions
    direction_colors = {
        'forward': 'rgba(93, 164, 214, 0.7)',   # Blue for forward
        'backward': 'rgba(255, 144, 14, 0.7)'   # Orange for backward
    }
    
    # Get unique directions in the data
    directions = data['direction'].unique()
    
    # Add boxplots for each category and direction combination
    for category in orderc:
        for direction in directions:
            # Get data for this category and direction
            category_direction_data = data[
                (data[var_x] == category) & 
                (data['direction'] == direction)
            ][var_name_y].dropna()
            
            if not category_direction_data.empty:
                # Get counts and statistics
                data_count = data_counts.get((category, direction), 0)
                trash_count = trash_counts.get((category, direction), 0)
                median = category_direction_data.median()
                mean = category_direction_data.mean()
                
                # Format category name with direction and count
                if trash_count == 0:
                    category_name = f"{category} ({direction}, n={data_count})"
                else:
                    category_name = f"{category} ({direction}, {data_count}/{data_count + trash_count})"
                
                # Add boxplot with direction-specific styling
                fig.add_trace(go.Box(
                    y=category_direction_data,
                    name=category_name,
                    legendgroup=direction,  # Group by direction for legend
                    legendgrouptitle_text=direction.capitalize(),
                    boxpoints='all',     # Show all points
                    pointpos=0,          # Center points horizontally
                    jitter=0.5,          # Add jitter to points
                    whiskerwidth=0.4,    # Thicker whiskers
                    marker=dict(
                        size=5,          # Slightly larger points
                        opacity=0.7,     # Semi-transparent
                        color='rgba(0,0,0,0.7)'  # Points in dark gray
                    ),
                    line=dict(width=1.5),  # Slightly thicker lines
                    fillcolor=direction_colors.get(direction, 'rgba(128, 128, 128, 0.7)'),
                    boxmean=True,        # Show mean
                    hoverinfo='all',
                    hovertemplate=(
                        f"<b>{category} ({direction})</b><br>" +
                        "Value: %{y:.3f}<br>" +
                        f"Median: {median:.3f}<br>" +
                        f"Mean: {mean:.3f}<br>" +
                        f"Count: {data_count}"
                    ),
                    # Add notches for visual appeal
                    notched=False,
                    notchwidth=0.5,
                    # Group boxes by category for proper pairing
                    offsetgroup=category,
                    x=[f"{category}"] * len(category_direction_data)  # Ensure proper x-axis positioning
                ))
    
    # Create a title with data information
    title_text = f"Paired Boxplot of {var_y} by {var_x} (Forward vs Backward)" + (" (filtered out)" if datatype == "junk" else "")
    subtitle = f"Data from {len(data)} ({trash.shape[0]} removed) measurements across {data[var_x].nunique()} {var_x} categories"
    
    # Update layout for paired boxplots
    fig.update_layout(
        title={
            'text': f"{title_text}<br><sup>{subtitle}</sup>",
            'y': 0.95,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top',
            'font': dict(size=18)
        },
        xaxis_title=var_x,
        yaxis_title=var_name_y,
        # Settings for paired boxplots
        boxmode='group',         # Group boxes by category
        boxgap=0.1,             # Gap between boxes within a group
        boxgroupgap=0.2,        # Gap between groups
        template="plotly_white",
        margin=dict(l=40, r=40, t=100, b=80),
        showlegend=True,        # Show legend for directions
        legend=dict(
            orientation="h",     # Horizontal legend
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        ),
        plot_bgcolor='rgb(243, 243, 243)',  # Light gray background
        paper_bgcolor='rgb(243, 243, 243)',
    )
    
    # Force the box width
    fig.update_traces(
        width=0.4,              # Narrower boxes to accommodate pairs
        quartilemethod="linear"  # Use linear method for quartiles
    )
    
    # Rotate x-axis labels if many categories
    if len(orderc) > 4:
        fig.update_layout(
            plot_bgcolor='white',    # Sets the plot area background to white
            paper_bgcolor='white',    # Sets the entire figure background to white
            xaxis=dict(
                tickangle=-10,
                tickfont=dict(size=10)
            )
        )
    
    # Save to Excel (using the same function as original)
    wb = save_combined_excel_data(path, wb, data, filtered_info, var_x, var_name_y, var_y, descriptor)

    if datatype == "junk":
        sample_name = f"boxplot_paired_j_{var_y}_by_{var_x}.html"
    else:
        sample_name = f"boxplot_paired_{var_y}_by_{var_x}.html"

    if not is_running_in_jupyter():
        fig.write_html(f"{path}{sample_name}")
        print(f"Saved paired boxplot of {var_y} by {var_x} (Forward vs Backward)")

    return fig, sample_name, wb


def save_full_data_frame(path, data):
    file_path = path + "0_numerical_results.xlsx"

    # Check if the Excel file already exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default sheet

    if not is_running_in_jupyter():
        with pd.ExcelWriter(file_path) as writer:
            # Write the DataFrame with earlier data to a sheet named 'Earlier'
            data.to_excel(writer, sheet_name='All_data')
    return wb


def save_combined_excel_data(path, wb, data, filtered_info, var_x, name_y, var_y, other_df):
    """Save data to Excel workbook - same functionality as the original but returning the workbook"""
    trash, filters = filtered_info
    # Define the Excel file path
    file_path = path + "0_numerical_results.xlsx"

    # Create a new sheet name based on var_x and var_y
    sheet_title = f"{var_y}-by-{var_x}"

    # Check if the sheet already exists
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    # Insert personalized string before the first DataFrame
    ws.append([f"Contents of boxplot for {var_y} by {var_x}"])
    ws.append([])  # Add an empty row for spacing

    # Process and append data and other_df as before
    combined_data = data.copy()
    combined_data['_index'] = combined_data.groupby(var_x).cumcount()
    pivot_table = combined_data.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        ws.append(r)

    # Calculate starting row for the second personalized string
    # It's the current number of rows plus 2 for spacing
    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    ws.cell(row=next_row, column=1, value="Statistical summary")
    ws.append([])  # Add an empty row for spacing

    for r in dataframe_to_rows(other_df.T, index=True, header=True):
        ws.append(r)

    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    ws.cell(row=next_row, column=1, value="This is the filtered data")
    ws.append([])  # Add an empty row for spacing

    combined_trash = trash.copy()
    combined_trash['_index'] = combined_trash.groupby(var_x).cumcount()
    pivot_table_trash = combined_trash.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    # Add rows from the second DataFrame (pivot table)
    for r in dataframe_to_rows(pivot_table_trash, index=True, header=True):
        ws.append(r)

    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    filter_words = ["Only data within these limits is shown:"] + filters
    for cc, strings in enumerate(filter_words):
        ws.cell(row=next_row + cc, column=1, value=strings)
    ws.append([])  # Add an empty row for spacing

    if not is_running_in_jupyter():
        # Save the workbook
        wb.save(filename=file_path)
    return wb


def center_plot(fig):
    """Center the Plotly plot on the data"""
    # Get all visible traces
    all_x = []
    all_y = []
    
    for trace in fig.data:
        # Skip if trace has no data
        if trace.x is None or len(trace.x) == 0 or trace.y is None or len(trace.y) == 0:
            continue
            
        # Skip legend-only traces
        if trace.get('showlegend', True) and (trace.x[0] is None or trace.y[0] is None):
            continue
            
        # Add x and y data to lists
        all_x.extend([x for x in trace.x if x is not None])
        all_y.extend([y for y in trace.y if y is not None])
    
    # If we have data, center the plot
    if all_x and all_y:
        x_min, x_max = min(all_x), max(all_x)
        y_min, y_max = min(all_y), max(all_y)
        
        # Add a small margin (10%)
        x_margin = (x_max - x_min) * 0.1
        y_margin = (y_max - y_min) * 0.1
        
        # Update the axis ranges
        fig.update_xaxes(range=[x_min - x_margin, x_max + x_margin])
        fig.update_yaxes(range=[y_min - y_margin, y_max + y_margin])
        
    return fig


def add_center_button(fig):
    """Add a custom button to center the plot"""
    fig.update_layout(
        updatemenus=[
            dict(
                type="buttons",
                direction="left",
                buttons=[
                    dict(
                        args=[{
                            "xaxis.autorange": False, 
                            "yaxis.autorange": False,
                            # Center function logic
                            "xaxis.range": [
                                min([min(trace.x) for trace in fig.data if trace.x and len(trace.x) > 0]) * 0.9,
                                max([max(trace.x) for trace in fig.data if trace.x and len(trace.x) > 0]) * 1.1
                            ],
                            "yaxis.range": [
                                min([min(trace.y) for trace in fig.data if trace.y and len(trace.y) > 0]) * 0.9,
                                max([max(trace.y) for trace in fig.data if trace.y and len(trace.y) > 0]) * 1.1
                            ]
                        }],
                        label="Center Plot",
                        method="relayout"
                    )
                ],
                pad={"r": 10, "t": 10},
                showactive=False,
                x=0.98,
                xanchor="right",
                y=0.02,
                yanchor="bottom"
            )
        ]
    )
    return fig


def histogram(path, df, var_y):
    """Create a histogram using Plotly"""
    names_dict = {
        'voc': 'Voc(V)', 'jsc': 'Jsc(mA/cm2)', 'ff': 'FF(%)', 'pce': 'PCE(%)',
        'vmpp': 'V_mpp(V)', 'jmpp': 'J_mpp(mA/cm2)', 'pmpp': 'P_mpp(mW/cm2)',
        'rser': 'R_series(Ohmcm2)', 'rshu': 'R_shunt(Ohmcm2)'
    }

    pl_y = names_dict[var_y]

    # Determine number of bins based on variable
    if var_y == "voc":
        bins = 20
    elif var_y == "jsc":
        bins = 30
    else:
        bins = 40

    # Create the histogram figure
    fig = go.Figure()
    
    # Add the histogram trace
    fig.add_trace(go.Histogram(
        x=df[pl_y],
        nbinsx=bins,
        marker=dict(
            color='rgba(0, 0, 255, 0.6)',
            line=dict(color='rgba(0, 0, 255, 1)', width=1)
        ),
        hovertemplate=f'{pl_y}: %{{x:.3f}}<br>Count: %{{y}}<extra></extra>'
    ))
    
    # Add a line trace for the kernel density estimate if enough data points
    if len(df) > 5:
        from scipy import stats
        kde_x = np.linspace(df[pl_y].min(), df[pl_y].max(), 100)
        kde = stats.gaussian_kde(df[pl_y].dropna())
        kde_y = kde(kde_x) * len(df) * (df[pl_y].max() - df[pl_y].min()) / bins
        
        fig.add_trace(go.Scatter(
            x=kde_x,
            y=kde_y,
            mode='lines',
            line=dict(color='red', width=2),
            name='KDE',
            hoverinfo='skip'
        ))
    
    # Add statistics annotations
    mean_val = df[pl_y].mean()
    median_val = df[pl_y].median()
    std_val = df[pl_y].std()
    
    stats_text = (
        f"Mean: {mean_val:.3f}<br>"
        f"Median: {median_val:.3f}<br>"
        f"Std Dev: {std_val:.3f}<br>"
        f"Count: {len(df)}"
    )
    
    fig.add_annotation(
        x=0.95,
        y=0.95,
        xref="paper",
        yref="paper",
        text=stats_text,
        showarrow=False,
        font=dict(size=12),
        align="right",
        bgcolor="rgba(255, 255, 255, 0.8)",
        bordercolor="black",
        borderwidth=1,
        borderpad=4
    )
    
    # Update layout
    fig.update_layout(
        title=f"Histogram of {pl_y}",
        xaxis_title=pl_y,
        yaxis_title="Frequency",
        template="plotly_white",
        bargap=0.1,
        hovermode='closest'
    )
    
    # Add grid lines
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')

    # Save figure if not in Jupyter
    sample_name = f"histogram_{var_y}.html"
    if not is_running_in_jupyter():
        fig.write_html(f"{path}{sample_name}")
        print(f"Saved histogram of {var_y}")

    return fig, sample_name


def condition_string_test(condition_var, unique_values):
    # is_condition = False
    if len(condition_var) > 1:
        condition_list = condition_var.replace(" ", "").split(',')
        if len(condition_list) == len(unique_values):
            is_condition = True
            return is_condition, condition_list
        else:
            print(
                f"Provided conditions: {len(condition_list)}. Number of samples requiring conditions: "
                f"{len(unique_values)}.")
            condition_string = input(
                f"Please re-enter, providing exactly {len(unique_values)} conditions for the samples, "
                f"each separated by a comma: ")

            condition_string_test(condition_string, unique_values)
    else:
        is_condition = False
        return is_condition, []


def plot_list_from_voila(plot_list):
    jvc_dict = {'Voc': 'v', 'Jsc': 'j', 'FF': 'f', 'PCE': 'p', 'R_ser': 'r', 'R_shu': 'h', 'V_mpp': 'u',
                'i_mpp': 'p', 'P_mpp': 'm'}
    box_dict = {'by Batch': 'e', 'by Variable': 'g', 'by Sample': 'a', 'by Cell': 'b', 'by Scan Direction': 'c'}
    cur_dict = {'All cells': 'Cy', 'Only working cells': 'Cz', 'Only not working cells': 'Co', 'Best device only': 'Cw',
                'Separated by cell': 'Cx', 'Separated by substrate': 'Cd'}

    new_list = []
    for plot in plot_list:
        code = ''
        if "omitted" in plot[0]:
            code += "J"
            code += (jvc_dict[plot[1]])
            code += (box_dict[plot[2]])
        elif "Boxplot" in plot[0]:
            code += "B"
            code += (jvc_dict[plot[1]])
            code += (box_dict[plot[2]])
        elif "Histogram" in plot[0]:
            code += "H"
            code += (jvc_dict[plot[1]])
        else:
            code += (cur_dict[plot[1]])
        new_list.append(code)

    return new_list


def plotting_string_action(plot_list, wb, data, supp, is_voila=False):
    filtered_jv, complete_jv, complete_cur = data
    omitted_jv, filter_pars, is_conditions, path, samples = supp

    if is_voila:
        plot_list = plot_list_from_voila(plot_list)

    # varplot_dict = {"B": "boxplot", "H": "Histogram", "C": "JV curve"}
    varx_dict = {"a": "sample", "b": "cell", "c": "direction", "d": "ilum", "e": "batch", "g": "condition", }
    vary_dict = {"v": "voc", "j": "jsc", "f": "ff", "p": "pce", "u": "vmpp", "i": "jmpp", "m": "pmpp", "r": "rser",
                 "h": "rshu", }
    # varc_dict = {"w": "best device", "x": "all cells per sample", "y": "all together"}

    fig_list = []
    fig_names = []
    for pl in plot_list:
        # Check if there is "condition":
        if "g" in pl and not is_conditions:
            continue
        # Check and assign var_x
        for key, value in varx_dict.items():
            if key in pl:
                var_x = value
                break  # Found var_x, no need to check further
        else:  # No break occurred
            var_x = None

        # Check and assign var_y
        for key, value in vary_dict.items():
            if key in pl:
                var_y = value
                break  # Found var_y, no need to check further
        else:  # No break occurred
            var_y = None

        # Check and plot varplot
        if "B" in pl and var_x is not None and var_y is not None:
            print(wb, var_x, var_y)
            fig, fig_name, wb = boxplot_all_cells(path, wb, filtered_jv, var_x, var_y, [omitted_jv, filter_pars], "data")
        elif "J" in pl and var_x is not None and var_y is not None:
            fig, fig_name, wb = boxplot_all_cells(path, wb, omitted_jv, var_x, var_y, [filtered_jv, filter_pars], "junk")
        elif "H" in pl and var_y is not None:
            fig, fig_name = histogram(path, complete_jv, var_y)
        elif "Cw" in pl:  # Best device
            fig, fig_name = jv_plot_curve_best(path, complete_jv, complete_cur)
        elif "Cx" in pl:  # Cells per sample
            for s in samples:
                fig, fig_name = jv_plot_by_cell_3x2(complete_cur, s, path)
                fig_list.append(fig)
                fig_names.append(fig_name)
            continue
        elif "Cd" in pl:  # Cells per substrate
            for s in samples:
                fig, fig_name = jv_plot_by_substrate(complete_cur, s, path)
                fig_list.append(fig)
                fig_names.append(fig_name)
            continue
        elif "Cy" in pl:  # All data
            #print("Plotting ALL data")
            fig, fig_name = jv_plot_together(complete_jv, complete_cur, path, "All")
        elif "Cz" in pl:  # Only filtered (working cells)
            #print("Plotting ONLY WORKING cells")
            fig, fig_name = jv_plot_together(filtered_jv, complete_cur, path, "Filtered")
        elif "Co" in pl:  # Only omitted (non-working cells)
            print(f"Plotting ONLY NON-WORKING cells")
            print(f"omitted_jv shape: {omitted_jv.shape}")
            print(f"sample/cell combinations: {omitted_jv[['sample', 'cell']].drop_duplicates().shape[0]}")
            fig, fig_name = jv_plot_together(omitted_jv, complete_cur, path, "Omitted")
        else:
            print(f"Command {pl} not recognized")
            continue

        fig_list.append(fig)
        fig_names.append(fig_name)

    return fig_list, fig_names, wb


def ask_to_input_initial_folder():
    user_path = input("Enter the path to the data folder for analysis. For example, C:\\Data\\Experiment : ")
    # user_path = r"D:\Seafile\JVData\Osail\20230717"
    path = user_path + '/'

    is_directory = os.path.isdir(path)

    if is_directory:
        listed_files = find_and_list_files(path)
        df_jvc, df_cur = load_files(listed_files)
        return df_jvc, df_cur, path
    else:
        print("Folder not found\n")
        return ask_to_input_initial_folder()


def find_unique_values(jvc_df):
    try:
        unique_values = jvc_df["identifier"].unique()
    except:
        unique_values = jvc_df["sample"].unique()
    print(f"\nThe following samples were found in the dataset: {', '.join(map(str, unique_values))}")

    return unique_values


def gather_conditions(unique_values):
    condition_var = input(
        "\nPress Enter to skip adding conditions. "
        "\nTo specify conditions for each sample, enter them no following the same order as above. "
        "Separate each condition with a comma. Leave a space for samples you wish to skip. "
        "Example: 1000 rpm, , 2500 rpm, 2500 rpm, 5000 rpm, ... :"
    )

    is_condition, list_conditions = condition_string_test(condition_var, unique_values)
    return is_condition, list_conditions


def gather_wanted_plots():
    plotting_string = input(
        "\nPress Enter to generate default plots: Boxplots (Voc, Jsc, FF, PCE) by sample, "
        "Histogram of PCE and all JV curves together."
        "\n\nTo create custom plots, enter codes as follows:"
        "\n  Plot Types: B=Boxplot, J=Boxplot(omitted), H=Histogram, C=JV curve"
        "\n  Parameters: a=sample, b=cell, c=direction, d=ilum, e=batch, g=condition"
        "\n              v=voc, j=jsc, f=ff, p=pce, u=vmpp, i=jmpp, m=pmpp, r=rser, h=rshu"
        "\n  JV Specific: Cw=best device only, Cx=only cells per sample, Cy=all data, Cz=only filtered, Co=only omitted"
        "\nExamples:"
        "\n  Bpg for a boxplot of PCEs by condition, Hv for a histogram of Voc,"
        "\n  Cy for all JV curves in a single plot."
        "\nWrite codes below separated by a comma, "
        "\nExample: Bfb, Hv, Cy : "
    )
    if len(plotting_string) > 1:
        plotting_list = plotting_string.replace(" ", "").split(',')
    else:
        plotting_list = ["Bav", "Baj", "Baf", "Bap", "Hp", "Cy"]

    return plotting_list


def create_new_results_folder(path):
    # Specify the path of the new folder
    folder_path = path + 'Results/'

    # Create the folder
    try:
        os.makedirs(folder_path)
    except FileExistsError:
        pass

    return folder_path


if __name__ == "__main__":
    # Individual actions
    jvc_data, cur_data, folder = ask_to_input_initial_folder()
    save_folder = create_new_results_folder(folder)
    workbook = save_full_data_frame(save_folder, jvc_data)

    if not jvc_data.empty:
        unique_vals = find_unique_values(jvc_data)
        is_cond, conditions = gather_conditions(unique_vals)
        jvc_data = name_by_condition(jvc_data, unique_vals, conditions)
        cur_data = name_by_condition(cur_data, unique_vals, conditions)
        jvc_filtered, junk, filter_vals = data_filter_setup(jvc_data, None)

        data_lists = [jvc_filtered, jvc_data, cur_data]
        extras = [junk, filter_vals, is_cond, save_folder, unique_vals]

        list_plots = gather_wanted_plots()
        # list_plots = ["Cz"]
        plotting_string_action(list_plots, workbook, data_lists, extras)

    print("Finished")